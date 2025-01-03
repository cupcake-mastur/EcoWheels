import html
import logging
import re
import stripe
import hashlib
import hmac
import os
import model
import random
import string
import secrets
import time
import requests


from flask import Flask, render_template, request, session, redirect, url_for, flash, current_app, jsonify, make_response, g, abort
from flask_wtf import CSRFProtect
from wtforms import StringField, TextAreaField, SelectField
from wtforms.validators import DataRequired, Email, AnyOf
from flask_mail import Mail, Message
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from Forms import CreateUserForm, UpdateProfileForm, LoginForm, OTPForm, RequestPasswordResetForm, ResetPasswordForm, AdminLoginForm, CreateVehicleForm, Feedbackform
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv, find_dotenv
from datetime import datetime, timedelta, timezone
from functools import wraps
from itsdangerous import URLSafeTimedSerializer
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import exists, func, text
from werkzeug.utils import secure_filename
from PIL import Image
from model import *
from flask_wtf.csrf import generate_csrf, CSRFError
from werkzeug.exceptions import BadRequest
from urllib.parse import unquote
import json
import qrcode
import pyotp
# ------------ For backup excel files -------------- #
from flask import send_file, jsonify
import pandas as pd
from io import BytesIO
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill
# ------------------------------------------------- #

load_dotenv(find_dotenv())
db = SQLAlchemy()

app = Flask(__name__)

logging.basicConfig(filename='app.log', level=logging.INFO,
                    format=f'%(asctime)s %(levelname)s: %(message)s')

app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB limit

app.config['SECRET_KEY'] = os.environ.get("SECRET_KEY")
s = URLSafeTimedSerializer(os.environ.get("SECRET_KEY"))
app.config['RECAPTCHA_PRIVATE_KEY'] = os.environ.get("RECAPTCHA_PRIVATE_KEY")
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get("DATABASE_URI")
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)  # Session timeout after 30 minutes

app.config.update(
    SESSION_COOKIE_SECURE=True,  # Only send cookie over HTTPS
    SESSION_COOKIE_HTTPONLY=True,  # Prevent JavaScript access to cookies
    SESSION_COOKIE_SAMESITE='Lax',  # Helps mitigate CSRF
)

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = os.getenv('EMAIL_USER')
app.config['MAIL_PASSWORD'] = os.getenv('EMAIL_PASS')
app.config['MAIL_DEFAULT_SENDER'] = os.getenv('EMAIL_USER')
mail = Mail(app)
otp_store = {}

user_logged_in = False
csrf = CSRFProtect(app)
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["10 per minute"]
)

with app.app_context():
    db.init_app(app)
    db.create_all()  # Create sql tables

SGT = pytz.timezone('Asia/Singapore')

#the stripe key for payment
stripe.api_key = os.getenv('STRIPE_SECRET_KEY')

# Retrieve the latest 10 payment intents
payment_intents = stripe.PaymentIntent.list(limit=10)

for intent in payment_intents.data:
    print(f"Payment Intent ID: {intent.id}, Amount: {intent.amount}, Status: {intent.status}")


def verify_recaptcha(response):
    # It expires in one minute**
    secret_key = app.config['RECAPTCHA_PRIVATE_KEY']
    payload = {'secret': secret_key, 'response': response}
    r = requests.post('https://www.google.com/recaptcha/api/siteverify', data=payload)
    return r.json().get('success', False)


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('user_logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


@limiter.request_filter
def exempt_routes():
    exempt_endpoints = [
        'dashboard', 'system_dashboard',
        'createVehicle', 'system_createVehicle', 'MCustomers', 'system_MCustomers', 'MVehicles',
        'system_MVehicles', 'system_logs', 'manageFeedback', 'system_manageFeedback',
        'sub_dashboard', 'sub_MCustomers', 'sub_MVehicles', 'sub_manageFeedback'
    ]
    return request.endpoint in exempt_endpoints


@app.errorhandler(429)
def ratelimit_error(e):
    app.logger.warning(f"Rate limit exceeded for IP {request.remote_addr}.")
    return render_template("customer/rate_limit_exceeded.html"), 429


@app.errorhandler(403)
def forbidden(e):
    return render_template('customer/403.html'), 403


@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    return render_template('error_msg.html', reason=e.description), 400


@app.route('/')
def home():
    return render_template("homepage/homepage.html")


# @app.route('/product_page')
# @login_required
# def product_page():
#     all_result = [
#         # Example product data (seller, product)
#         ('Seller1', {'get_product_id': lambda: 'prod_1', 'get_product_name': lambda: 'Product 1', 'get_product_price': lambda: 15.00, 'get_image': lambda: 'product1.jpg'}),
#         ('Seller2', {'get_product_id': lambda: 'prod_2', 'get_product_name': lambda: 'Product 2', 'get_product_price': lambda: 5.00, 'get_image': lambda: 'product2.jpg'}),
#         ('Seller3', {'get_product_id': lambda: 'prod_3', 'get_product_name': lambda: 'Product 3', 'get_product_price': lambda: 8.00, 'get_image': lambda: 'product3.jpg'}),
#         ('Seller4', {'get_product_id': lambda: 'prod_4', 'get_product_name': lambda: 'Product 4', 'get_product_price': lambda: 2.00, 'get_image': lambda: 'product4.jpg'}),
#     ]
#     return render_template('customer/test_product_page(exists till terron creates one hehe).html', all_result=all_result)


@app.route('/models')
@login_required
def models():
    vehicles = db.session.query(Vehicle).all()

    return render_template("homepage/models.html" , vehicles=vehicles)

@app.route('/afford_calc.html')
def afford_calc():
    return render_template('homepage/afford_calc.html')

@app.route('/used_car_calc.html')
def used_car_calc():
    return render_template('homepage/used_car_calc.html')



class FeedbackForm(FlaskForm):
    email = StringField('Email', validators=[DataRequired(), Email()])
    rating = SelectField('Rating', choices=[('good', 'Good'), ('moderate', 'Moderate'), ('bad', 'Bad')], validators=[DataRequired(), AnyOf(['good', 'moderate', 'bad'])])
    feedback = TextAreaField('Feedback', validators=[DataRequired()])


limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["20 per day", "30 per hour"]  # Adjust these limits as needed
)
@app.route('/Feedback', methods=['GET', 'POST'])
@login_required  # Ensure only logged-in users can access this route
@limiter.limit("20 per minute")
def feedback():
    form = FeedbackForm()
    user_email = session.get('user_email')
    user = db.session.query(User).filter_by(email=user_email).first()

    if not user:
        flash('You must be logged in to submit feedback.', 'error')
        return redirect(url_for('login'))

    username = user.username
    user_id = user.id  # Ensure user_id is always set

    if request.method == 'POST':
        if form.validate_on_submit():
            email = form.email.data
            rating = form.rating.data
            feedback_message = form.feedback.data

            try:
                new_feedback = Feedback(
                    username=username,
                    email=email,
                    feedback=feedback_message,
                    rating=rating,
                    user_id=user_id  # Enforce user_id
                )
                db.session.add(new_feedback)
                db.session.commit()
                flash('Thank you for your feedback!', 'success')
                return redirect(url_for('thankyou'))

            except Exception as e:
                db.session.rollback()
                flash(f'An error occurred. Please try again later. Error: {str(e)}', 'error')
        else:
            flash('Please correct the errors and try again.', 'error')

    return render_template('homepage/Feedback.html', username=username, form=form, user_id=user_id)

def admin_login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_log_in'))  # Redirect to admin login if not logged in
        return f(*args, **kwargs)
    return decorated_function

# Admin Session time-out
def admin_session_timeout_check(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        now = datetime.now().timestamp()
        last_activity = session.get('admin_last_activity', None)

        if last_activity and now - last_activity > 15 * 60:  # 15 minutes timeout
            admin_username = session.get('admin_username')
            session.pop('admin_logged_in', None)
            session.pop('admin_username', None)
            session.pop('admin_role', None)
            session.pop('admin_last_activity', None)
            session.pop('security_modal_shown', None)
            log_event("Logout", f"Inactivity for 15 minutes led to {admin_username}'s session timeout")
            return redirect(url_for('admin_log_in'))

        session['admin_last_activity'] = now
        return f(*args, **kwargs)
    return decorated_function

def is_valid_input(input_str):
    """
    Check if the input string contains only allowed characters.
    """
    # Define a regular expression to match allowed characters
    allowed_chars_pattern = re.compile(r'^[\w.@+-]+$')
    return bool(allowed_chars_pattern.match(input_str))

def role_required(*roles):
    def wrapper(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'admin_logged_in' not in session:
                return redirect(url_for('admin_log_in'))
            if session.get('admin_role') not in roles:
                return redirect(url_for('ErrorPage'))
            return f(*args, **kwargs)
        return decorated_function
    return wrapper

@app.route('/manageFeedback', methods=['GET', 'POST'])
@admin_login_required
@admin_session_timeout_check
@role_required('general')
def manageFeedback():
    admin_username = session.get('admin_username')
    user_email = session.get('user_email')
    csrf_token = generate_csrf()  # Generate CSRF token
    accepted_ratings = ('good', 'moderate', 'bad')
    errors = {}

    if request.method == 'POST':
        # Get filter criteria from the form
        customer_id = request.form.get('customer_id')
        email = request.form.get('email')
        rating = request.form.get('rating')

        # Build the query with filters
        query = db.session.query(Feedback)

        if customer_id and is_valid_input(customer_id) and customer_id.isdigit():
            query = query.filter(Feedback.user_id == int(customer_id))
        elif customer_id:
            errors['customer_id'] = "Invalid input for customer id"

        if email and is_valid_input(email):
            query = query.filter(Feedback.email.ilike(f"%{email}%"))
        elif email:
            errors['email'] = "Invalid input for email"

        if rating and rating in accepted_ratings:
            query = query.filter(Feedback.rating == rating)
        elif rating:
            errors['rating'] = "Invalid rating"

        # Get all feedback entries based on filters
        feedback_entries = query.all()
    else:
        # Query the Feedback table to get all feedback entries
        feedback_entries = db.session.query(Feedback).all()

    # Render the template with the feedback entries
    return render_template('admin/manageFeedback.html', admin_username=admin_username,
                           feedback_entries=feedback_entries, csrf_token=csrf_token, errors=errors)
@app.route('/sub_manageFeedback', methods=['GET', 'POST'])
@admin_login_required
@admin_session_timeout_check
@role_required('junior')
def sub_manageFeedback():
    admin_username = session.get('admin_username')
    csrf_token = generate_csrf()  # Generate CSRF token
    accepted_ratings = ('good', 'moderate', 'bad')
    errors = {}

    if request.method == 'POST':
        # Get filter criteria from the form
        customer_id = request.form.get('customer_id')
        email = request.form.get('email')
        rating = request.form.get('rating')

        # Build the query with filters
        query = db.session.query(Feedback)

        if customer_id and is_valid_input(customer_id) and customer_id.isdigit():
            query = query.filter(Feedback.user_id == int(customer_id))
        elif customer_id:
            errors['customer_id'] = "Invalid input for customer id"

        if email and is_valid_input(email):
            query = query.filter(Feedback.email.ilike(f"%{email}%"))
        elif email:
            errors['email'] = "Invalid input for email"

        if rating and rating in accepted_ratings:
            query = query.filter(Feedback.rating == rating)
        elif rating:
            errors['rating'] = "Invalid rating"

        # Get all feedback entries based on filters
        feedback_entries = query.all()
    else:
        # Query the Feedback table to get all feedback entries
        feedback_entries = db.session.query(Feedback).all()

    # Render the template with the feedback entries
    return render_template('admin/junior_admin/sub_manageFeedback.html', admin_username=admin_username,
                           feedback_entries=feedback_entries, errors=errors, csrf_token=csrf_token)


@app.route('/delete_feedback/<int:feedback_id>', methods=['POST'])
@admin_login_required
def delete_feedback(feedback_id):  # Include feedback_id in the function signature
    feedback = db.session.query(Feedback).get(feedback_id)  # Query the specific feedback entry by ID
    if feedback:
        db.session.delete(feedback)  # Delete the specific feedback entry
        db.session.commit()
    if session['admin_role'] == 'system':
        return redirect(url_for('system_manageFeedback'))
    elif session['admin_role'] == 'general':
        return redirect(url_for('manageFeedback'))
    else:
        return redirect(url_for('ErrorPage'))


@app.route('/thankyou')
def thankyou():
    return render_template("homepage/thankyou.html")


@app.route('/visit', methods=['POST'])
@limiter.limit("20 per minute")
@csrf.exempt
@login_required
def visit():
    EXEMPT_URL_ROUTES = ['http://127.0.0.1:5000/sign_up', 'http://127.0.0.1:5000/login', 'http://127.0.0.1:5000/verify_otp']
    data = request.json
    url = data.get('url')
    email = session.get('user_email')
    user_id = get_user_id_from_email(email)

    if url in EXEMPT_URL_ROUTES:
        return jsonify({'message': 'URL is exempt from recording'})

    # Ensure user_id is not None
    if user_id is None:
        raise ValueError("User ID is None")

    visited_at = datetime.now()

    # Remove any previous entries for this user
    db.session.query(UserURL).filter_by(user_id=user_id).delete()
    db.session.commit()

    # Insert the new entry
    new_entry = UserURL(email=email, user_id=user_id, url=url, visited_at=visited_at)
    db.session.add(new_entry)
    db.session.commit()

    message = 'Latest URL visit recorded'

    return jsonify({'message': message, 'url': url}), 200

def get_user_id_from_email(email):
    user = db.session.query(User).filter_by(email=email).first()
    return user.id if user else None


@app.route('/check_session')
@limiter.limit("20 per minute")
def check_session():
    # Log the current state of the session for debugging
    app.logger.info("Session data: %s", session)

    if 'expiry_time' in session:
        current_time = datetime.now(timezone.utc).timestamp()
        expiry_time = session['expiry_time']

        app.logger.info("Current time: %s", current_time)
        app.logger.info("Session expiry time: %s", expiry_time)

        if current_time > expiry_time:
            app.logger.info("Session expired: True")
            return jsonify(expired=True), 200

    app.logger.info("Session expired: False")
    return jsonify(expired=False), 200


def hash_password_sha1(password):
    """Hashes the password using SHA-1."""
    sha1 = hashlib.sha1()
    sha1.update(password.encode('utf-8'))
    return sha1.hexdigest().upper()

def check_pwned_password(hashed_password):
    """Checks the hashed password against the HIBP API."""
    prefix = hashed_password[:5]  # First 5 characters of the hashed password
    suffix = hashed_password[5:]  # Remaining characters
    url = f"https://api.pwnedpasswords.com/range/{prefix}"
    response = requests.get(url)
    return suffix, response.text

def is_password_pwned(suffix, response_text):
    """Compares the password suffix with the API response to check if it has been pwned."""
    lines = response_text.splitlines()
    for line in lines:
        hash_suffix, count = line.split(":")
        if suffix == hash_suffix:
            return True, count
    return False, 0


@app.route('/sign_up', methods=['GET', 'POST'])
def sign_up():
    error = None
    create_user_form = CreateUserForm(request.form)
    if request.method == 'POST' and create_user_form.validate():
        full_name = create_user_form.full_name.data
        username = create_user_form.username.data
        email = create_user_form.email.data
        phone_number = create_user_form.phone_number.data
        password = create_user_form.password.data
        confirm_password = create_user_form.confirm_password.data

        user_exists = db.session.query(exists().where(User.username == username)).scalar()
        email_exists = db.session.query(exists().where(User.email == email)).scalar()
        phone_number_exists = db.session.query(exists().where(User.phone_number == phone_number)).scalar()

        if user_exists:
            error = "Username already exists!"
        elif email_exists:
            error = "Email is already registered!"
        elif phone_number_exists:
            error = "Phone number is already registered!"
        elif password != confirm_password:
            error = "Passwords do not match!"
        elif len(str(phone_number)) != 8:
            error = "Phone number must be 8 digits."
        else:
            special_chars = "!@#$%^&*()-_=+[{]}\\|;:'\",<.>/?"

            if any(char in special_chars for char in full_name) or any(char in special_chars for char in username):
                error = "Special characters are not allowed in the full name or username."
            elif not any(char in special_chars for char in password):
                error = "Password must contain at least one special character."  
            elif not any(char.isupper() for char in password) or not any(char.islower() for char in password):
                error = "Password must contain at least one uppercase and lowercase letter."

        if error is None:
            # Generate the SHA-1 hash of the password for HIBP check
            hashed_password_sha1 = hash_password_sha1(password)
            
            try:
                # Check if the password is compromised
                suffix, response_text = check_pwned_password(hashed_password_sha1)
                is_pwned, count = is_password_pwned(suffix, response_text)
                
                if is_pwned:
                    error = f"Your password has been detected in previous data breaches.\n For your security, please select a different password."

                else:
                    hashed_password = generate_password_hash(password)
                    print(f"Generated password hash: {hashed_password}") 

                    new_user = User(full_name=full_name, username=username, email=email, phone_number=phone_number,
                                    password_hash=hashed_password)
                    db.session.add(new_user)
                    db.session.commit()

                    password_history = PasswordHistory(user_id=new_user.id, password_hash=hashed_password, changed_at=datetime.now(timezone.utc))
                    db.session.add(password_history)
                    db.session.commit()

                    app.logger.info(f"User {email} added to database.")
                    return redirect(url_for('login'))
            except requests.RequestException as e:
                app.logger.error(f"Error checking password with HIBP: {e}")
                error = "There was an error checking your password. Please try again later."
    return render_template("customer/sign_up.html", form=create_user_form, error=error)


def generate_otp(length=6):
    otp = ''.join(random.choices(string.digits, k=length))
    session['otp_generation_time'] = datetime.now(timezone.utc).timestamp()
    return otp


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    print(session)
    if session.get('user_logged_in'):
        return render_template('customer/error_msg_already_logged_in.html')

    login_form = LoginForm(request.form)

    if request.method == 'POST' and login_form.validate():
        email = html.escape(login_form.email.data)
        password = html.escape(login_form.password.data)
        user = db.session.query(User).filter_by(email=email).first()
        recaptcha_response = request.form.get('g-recaptcha-response')
        
        # Verify the reCAPTCHA response
        if not verify_recaptcha(recaptcha_response):
            error = "reCAPTCHA verification failed. Please try again."
            app.logger.warning(f"Failed reCAPTCHA verification for {email}")
            return render_template("customer/login.html", form=login_form, error=error)

        if user:
            print(f"Stored password hash: {user.password_hash}")  
            print(f"Entered password: {password}")  
            current_time = datetime.now(SGT)

            if user.lockout_until and user.lockout_until.tzinfo is None:
                user.lockout_until = SGT.localize(user.lockout_until)

            if user.lockout_count >= 4:
                error = "Account is permanently locked. Please contact us for assistance."
                app.logger.warning(f"Attempted login for permanently locked account {email}")

            elif user.lockout_until and user.lockout_until > current_time:
                error = "Account is locked. Please try again later."
                app.logger.warning(f"Locked account login attempt for {email}")
            else:
                # If lockout_until has expired, reset failed attempts and lockout_until
                if user.lockout_until and user.lockout_until <= current_time:
                    user.failed_attempts = 0
                    user.lockout_until = None
                    db.session.commit()

                if check_password_hash(user.password_hash, password):
                    print("Password matched!") 
                    user.failed_attempts = 0
                    user.lockout_until = None
                    db.session.commit()

                    otp = generate_otp()
                    session['otp'] = otp
                    session['unverified_user_email'] = email

                    send_otp_email(user.email, otp)
                    app.logger.info(f"OTP sent to {user.email}")

                    return redirect(url_for('verify_otp'))
                else:
                    print("Password did not match.")
                    user.failed_attempts += 1

                    if user.failed_attempts >= 3:
                        user.lockout_count += 1
                        lockout_duration = get_lockout_duration(user.lockout_count)
                        if lockout_duration:
                            user.lockout_until = current_time + lockout_duration
                            error = f"Too many failed attempts. Account is locked for {lockout_duration}."
                            app.logger.warning(f"Account locked for {email} after {user.failed_attempts} failed attempts.")
                        else:
                            user.lockout_until = None  # Permanent lockout
                            error = "Account is permanently locked. Please contact us for assistance."
                            app.logger.warning(f"Account permanently locked for {email} after {user.lockout_count} lockouts.")
                    else:
                        error = "Invalid email or password. Please try again."
                        app.logger.warning(f"Failed login attempt for {email}")

                    db.session.commit()
        else:
            error = "Invalid email or password. Please try again."
            app.logger.warning(f"Failed login attempt for {email}")

    return render_template("customer/login.html", form=login_form, error=error)


def get_lockout_duration(lockout_count):
    if lockout_count == 1:
        return timedelta(seconds=30) 
    elif lockout_count == 2:
        return timedelta(seconds=40)
    elif lockout_count == 3:
        return timedelta(seconds=50)
    else:
        return None  # Permanent lockout


def send_otp_email(email, otp):
    msg = Message('Your OTP Code', recipients=[email])
    msg.body = f'Your OTP code is: {otp}\n\nPlease note that this code will expire in 1 minute.'
    mail.send(msg)


@app.route('/request_new_otp', methods=['POST'])
def request_new_otp():
    user_email = session.get('user_email')
    session.pop('otp', None)
    new_otp = generate_otp()
    session['otp'] = new_otp
    send_otp_email(user_email, new_otp)
    app.logger.info(f"New OTP sent to {user_email}")

    return redirect(url_for('verify_otp'))


def hide_email(email):
    parts = email.split('@')
    return parts[0][:2] + '****' + parts[0][-2:] + '@' + parts[1]


app.jinja_env.filters['hide_email'] = hide_email


@app.route('/verify_otp', methods=['GET', 'POST'])
def verify_otp():
    error = None
    otp_form = OTPForm(request.form)

    user_email = session.get('unverified_user_email')  

    if not user_email:
        return redirect(url_for('login'))

    user = db.session.query(User).filter_by(email=user_email).first()
    if not user:
        return redirect(url_for('login'))

    if request.method == 'POST' and otp_form.validate():
        entered_otp_digits = [
            otp_form.otp1.data,
            otp_form.otp2.data,
            otp_form.otp3.data,
            otp_form.otp4.data,
            otp_form.otp5.data,
            otp_form.otp6.data
        ]
        entered_otp = ''.join(entered_otp_digits)
        print(entered_otp)
        otp = session.get('otp')
        otp_generation_time = session.get('otp_generation_time')
        current_time = datetime.now(timezone.utc).timestamp()

        if otp_generation_time and (current_time - otp_generation_time) <= 60:
            if entered_otp == otp:
                session.pop('otp', None)
                session.pop('otp_generation_time', None)
                session.pop('unverified_user_email', None)

                session['user_email'] = user_email 
                session['expiry_time'] = (datetime.now(timezone.utc) + app.config['PERMANENT_SESSION_LIFETIME']).timestamp()
                session['user_logged_in'] = True
                session.permanent = True

                last_visited = db.session.query(UserURL).filter_by(user_id=user.id).first()
                if last_visited:
                    app.logger.info(f"User {user_email} logged in successfully.")
                    return redirect(last_visited.url)
                else:
                    return redirect(url_for('home'))
            else:
                error = "Invalid OTP. Please try again."
                app.logger.warning(f"Invalid OTP attempt for {user_email}")
        else:
            error = "Invalid OTP. Please try again."
            app.logger.warning(f"Expired OTP attempt for {user_email}")

    return render_template('customer/verify_otp.html', form=otp_form, error=error, user_email=user_email)


@app.route('/profile', methods=['GET'])
@login_required
def profile():
    user_email = session.get('user_email')

    user = db.session.query(User).filter_by(email=user_email).first()

    if not user_email or not user:
        return redirect(url_for('login'))

    return render_template('customer/profile_page.html', user=user)


@app.route('/request_password_reset', methods=['GET', 'POST'])
def request_password_reset():
    error = None
    request_password_reset_form = RequestPasswordResetForm(request.form)

    if request.method == 'POST' and request_password_reset_form.validate():
        email = request.form.get('email')
        user = db.session.query(User).filter_by(email=email).first()

        recent_reset_requests = db.session.query(PasswordResetRequest).filter(
            PasswordResetRequest.last_request_time >= datetime.now(SGT) - timedelta(seconds=40)
        ).all()

        total_requests = sum(req.request_count for req in recent_reset_requests)

        if total_requests >= 3:
            error = (
                "Password reset request limit exceeded.\n" 
                "Please try again later."
            )
            return render_template('customer/request_password_reset.html', form=request_password_reset_form, error=error)

        reset_request = db.session.query(PasswordResetRequest).filter_by(email=email).first()
        if not reset_request:
            reset_request = PasswordResetRequest(email=email)
            db.session.add(reset_request)

        # Record the request
        if not reset_request.can_request():
            error = (
                "Password reset request limit exceeded for this email.\n" 
                "Please try again later."
            )
            return render_template('customer/request_password_reset.html', form=request_password_reset_form, error=error)

        reset_request.record_request()
        db.session.commit()

        # Send reset email if user exists
        user = db.session.query(User).filter_by(email=email).first()
        if user:
            reset_request.user_id = user.id
            db.session.commit()

            token = s.dumps(email, salt='password-reset-salt')
            reset_url = url_for('reset_password', token=token, _external=True)
            send_password_reset_email(email, reset_url)
            app.logger.info(f"Reset Link has been sent to {email} successfully.")
            return render_template('customer/request_password_success.html')
        else:
            app.logger.warning(f"Invalid email {email} entered.")
            return render_template('customer/request_password_success.html')

    return render_template('customer/request_password_reset.html', form=request_password_reset_form, error=error)


def send_password_reset_email(email, reset_url):
    msg = Message('Password Reset Request', recipients=[email])
    msg.html = f'''
    <p>Click the link below to reset your password:</p>
    <p><a href="{reset_url}">Reset Password</a></p>
    <p>If you did not request this, please contact us at +6562911189.</p>
    '''
    mail.send(msg)


@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    error = None
    reset_password_form = ResetPasswordForm(request.form)
    email = None
    try:
        email = s.loads(token, salt='password-reset-salt', max_age=60)
    except Exception as e:
        app.logger.warning(f"Reset Link is invalid or expired: {e}")
        return render_template('customer/invalid_token.html')

    if request.method == 'POST' and reset_password_form.validate():
        if email is None:
            return render_template('customer/invalid_token.html')
        
        new_password = reset_password_form.password.data
        confirm_password = reset_password_form.confirm_password.data

        special_chars = "!@#$%^&*()-_=+[{]}\\|;:'\",<.>/?"

        if new_password != confirm_password:
            error = "Passwords do not match."
        elif not any(char in special_chars for char in new_password):
            error = "Password must contain at least one special character."
        elif not any(char.isupper() for char in new_password) or not any(char.islower() for char in new_password):
            error = "Password must contain at least one uppercase and one lowercase letter."
        if error:
            print(f"Error detected: {error}")
            return render_template('customer/reset_password.html', form=reset_password_form, token=token, error=error)

        user = db.session.query(User).filter_by(email=email).first()
        if user:
            user.password_hash = generate_password_hash(new_password)
            new_password_history = PasswordHistory(user_id=user.id, password_hash=user.password_hash)
            db.session.add(new_password_history)

            all_passwords = (db.session.query(PasswordHistory).filter_by(user_id=user.id)
                             .order_by(PasswordHistory.changed_at.desc()).all())
            if len(all_passwords) > 3:
                for old_password in all_passwords[3:]:
                    db.session.delete(old_password)
                    
            db.session.commit()
            return render_template('customer/password_reset_success.html')
        else:
            error = "An error occurred. Please try again."
            app.logger.error(f"User with email {email} not found.")

    return render_template('customer/reset_password.html', form=reset_password_form, token=token, error=error)


@app.route('/edit_profile', methods=['GET', 'POST'])
@login_required
def edit_profile():
    error = None
    user_email = session.get('user_email')
    user = db.session.query(User).filter_by(email=user_email).first()

    if not user:
        return redirect(url_for('login'))

    edit_profile_form = UpdateProfileForm(request.form, obj=user)

    if request.method == 'POST' and edit_profile_form.validate():
        full_name = edit_profile_form.full_name.data
        username = edit_profile_form.username.data
        email = edit_profile_form.email.data
        phone_number = edit_profile_form.phone_number.data
        current_password = edit_profile_form.current_password.data
        new_password = edit_profile_form.new_password.data
        confirm_new_password = edit_profile_form.confirm_new_password.data

        if phone_number != user.phone_number:
            if len(str(phone_number)) != 8:
                error = "Phone number must be 8 digits."
            else:
                phone_number_exists = db.session.query(User).filter(User.phone_number == phone_number, User.id != user.id).first()
                if phone_number_exists:
                    error = "Phone number is already registered."

        if email != user.email:
            email_exists = db.session.query(User).filter(User.email == email, User.id != user.id).first()
            if email_exists:
                error = "Email is already registered."

        if current_password and not check_password_hash(user.password_hash, current_password):
            error = 'Current password is incorrect.'
        elif new_password and new_password != confirm_new_password:
            error = 'New passwords do not match.'
        else:
            special_chars = "!@#$%^&*()-_=+[{]}\\|;:'\",<.>/?"
            if any(char in special_chars for char in full_name) or any(char in special_chars for char in username):
                error = "Special characters are not allowed in the full name or username."
            elif new_password:
                if not any(char in special_chars for char in new_password):
                    error = "Password must contain at least one special character."
                elif not any(char.isupper() for char in new_password) or not any(char.islower() for char in new_password):
                    error = "Password must contain at least one uppercase and lowercase letter."

            # Check last 3 passwords
            if not error:
                recent_passwords = db.session.query(PasswordHistory).filter_by(user_id=user.id).order_by(PasswordHistory.changed_at.desc()).limit(3).all()
                for past_password in recent_passwords:
                    if check_password_hash(past_password.password_hash, new_password):
                        error = 'New password cannot be the same as any of the last 3 passwords.'
                        break

            if not error:
                user.full_name = full_name
                user.username = username
                user.email = email
                user.phone_number = phone_number

                # Update password if new password is provided and matches confirmation
                if new_password:
                    user.password_hash = generate_password_hash(new_password)
                    new_password_history = PasswordHistory(user_id=user.id, password_hash=user.password_hash)
                    db.session.add(new_password_history)

                    # Keep only the last 3 password hashes
                    # Most recent password changes are listed first (because of the desc)
                    all_passwords = db.session.query(PasswordHistory).filter_by(user_id=user.id).order_by(PasswordHistory.changed_at.desc()).all()
                    if len(all_passwords) > 3:
                        for old_password in all_passwords[3:]:
                            db.session.delete(old_password)

        if not error:
            db.session.commit()
            error = 'Profile updated successfully.'

    return render_template('customer/edit_profile.html', user=user, form=edit_profile_form, error=error)


@app.route('/user/logout')
def logout():
    if 'user_email' in session:
        user_email = session.get('user_email')
        session.clear()
        session.modified = True
        app.logger.info(f"User {user_email} logged out successfully.")
        print(session)
        return render_template('customer/logout_message.html')
    else:
        return render_template('customer/error_msg_not_logged_in.html')


@app.route('/delete_account', methods=['GET', 'POST'])
def delete_account():    
    user_email = session.get('user_email')
    if user_email:
        user = db.session.query(User).filter_by(email=user_email).first()
        if user:
            user_id = user.id
            # Delete entries from related tables first (because of foreign key constraints)
            db.session.query(UserURL).filter_by(user_id=user_id).delete()
            db.session.query(PasswordHistory).filter_by(user_id=user_id).delete()
            db.session.query(PasswordResetRequest).filter_by(user_id=user.id).delete()
            db.session.delete(user)
            db.session.commit()
            session.clear()
            session.modified = True
            return render_template('customer/account_deleted_successfully.html')
        else:
            return render_template('error_msg.html')
    else:
        return render_template('customer/error_msg_not_logged_in.html')


@app.before_request
def before_request():
    if 'user_email' in session:
        current_time = datetime.now(timezone.utc).timestamp()
        if 'expiry_time' in session and current_time > session['expiry_time']:
            session.clear()
            session.modified = True
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify(expired=True), 401  # Return JSON for the AJAX request

        if request.headers.get('X-Check-Session') != 'True':
            session['expiry_time'] = (datetime.now(timezone.utc) + app.config['PERMANENT_SESSION_LIFETIME']).timestamp()
            session.modified = True


# @app.route('/cart')
# def cart_page():
#     return render_template('customer/shopping_cart.html')

def get_latest_payment_intents(limit=10):
    payment_intents = stripe.PaymentIntent.list(limit=limit)
    return payment_intents['data']

@app.route('/stripe_dashboard')
def stripe_dashboard():
    payment_intents = get_latest_payment_intents()
    return render_template('admin/stripe_dashboard.html', payment_intents=payment_intents)


@app.route('/create-checkout-session', methods=['POST'])
@login_required
def create_checkout_session():
    try:
        data = request.get_json()
        product_id = data['product_id']
        model = data['model']
        amount = data['amount']

        email = session['user_email']
        user = db.session.query(User).filter_by(email=email).first()

        if not user:
            return jsonify(error='User not found'), 404

        customer_name = user.full_name
        customer_email = user.email

        # Create a customer in Stripe if necessary
        stripe_customer = stripe.Customer.create(
            name=customer_name,
            email=customer_email
        )

        checkout_session = stripe.checkout.Session.create(
            payment_method_types=['card'],
            customer=stripe_customer.id,
            line_items=[{
                'price_data': {
                    'currency': 'usd',
                    'product_data': {
                        'name': model,
                        'metadata': {
                            'product_id': product_id,
                            'model': model,
                        }
                    },
                    'unit_amount': int(amount * 100),
                },
                'quantity': 1,
            }],
            mode='payment',
            success_url='http://127.0.0.1:5000/payment_success',
            cancel_url='http://127.0.0.1:5000/cancel',
        )
        return jsonify({'id': checkout_session.id})
    except Exception as e:
        return jsonify(error=str(e)), 400
        
def payment_access_control(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user_id = session.get('user_id')
        
        # Query to check if the current session user is authorized to access the payment process
        if not user_id or not is_user_authorized_for_payment(user_id):
            return abort(403)  # Deny access if the check fails
        
        return f(*args, **kwargs)
    return decorated_function

def is_user_authorized_for_payment(user_id):
    # Implement your logic to verify if the user is authorized
    # For example, check if the user has a valid session or matches criteria for accessing payment resources
    user = db.session.query(User).filter_by(id=user_id).first()
    return user is not None  # Modify this according to your requirements

# @app.route('/customer/history')
# def customer_purchase_history():
#     # Execute the query to fetch all purchases
#     purchases = db.session.execute(text("""
#         SELECT full_name, email, product_name, price 
#         FROM PurchasedItem
#         ORDER BY email;
#     """)).fetchall()

#     # Dictionary to group products by customer email
#     customer_purchases = {}

#     for purchase in purchases:
#         email = purchase.email
#         if email not in customer_purchases:
#             customer_purchases[email] = {
#                 'full_name': purchase.full_name,
#                 'email': email,
#                 'products': []
#             }
#         customer_purchases[email]['products'].append({
#             'prod_name': purchase.product_name,
#             'price': purchase.price
#         })

#     # Render the template with grouped customer purchases
#     return render_template('admin/customer_purchase_history.html', customers=customer_purchases)



# @app.route('/purchased-items')
# def purchased_items():
#     products =  db.session.query(Product).all()
#     return render_template('admin/purchased_items.html', products=products)
# # def purchased_items():
#     payment_intents = stripe.PaymentIntent.list(limit=10)
#     purchased_items = []
#     for intent in payment_intents.data:
#         if intent.status == 'succeeded':
#             customer_name = 'N/A'
#             customer_email = 'N/A'
#             product_id = intent.metadata.get('product_id', 'N/A')
            
#             if intent.customer:
#                 customer = stripe.Customer.retrieve(intent.customer)
#                 customer_name = customer.name
#                 customer_email = customer.email
            
#             item = {
#                 'customer_name': customer_name,
#                 'customer_email': customer_email,
#                 'product_id': product_id,
#                 'amount': intent.amount / 100,
#                 'currency': intent.currency.upper(),
#                 'purchase_date': datetime.fromtimestamp(intent.created)
#             }
#             purchased_items.append(item)
#     return render_template('admin/purchased_items.html', purchased_items=purchased_items)

# @app.route('/fetch-purchases', methods=['GET'])
# def fetch_purchases():
#     try:
#         payment_intents = stripe.PaymentIntent.list(limit=10)
#         for intent in payment_intents.data:
#             if intent.status == 'succeeded':
#                 customer_name = intent.charges.data[0].billing_details.name
#                 customer_email = intent.charges.data[0].billing_details.email
#                 product_id = intent.metadata.get('product_id')
#                 model = intent.metadata.get('model')
#                 amount = intent.amount_received / 100.0  # Stripe amounts are in cents
#                 currency = intent.currency.upper()
#                 purchase_date = datetime.fromtimestamp(intent.created, tz=SGT)
                
#                 # Check if the purchase already exists
#                 existing_purchase = db.session.query(Purchase).filter_by(payment_intent_id=intent.id).first()
#                 if not existing_purchase:
#                     new_purchase = Purchase(
#                         customer_name=customer_name,
#                         customer_email=customer_email,
#                         product_id=product_id,
#                         model=model,
#                         amount=amount,
#                         currency=currency,
#                         purchase_date=purchase_date
#                     )
#                     db.session.add(new_purchase)
#                     db.session.commit()
                    
#         return jsonify({'status': 'success'}), 200
#     except Exception as e:
#         return jsonify(error=str(e)), 400


@app.route('/payment_success')
def success_page():
    return render_template('customer/payment_success.html')

@app.route('/cancel')
def cancel_page():
    return "Payment Cancelled"
# @app.route('/payment')
# def payment():
#     return render_template("customer/payment.html")


# @app.route('/confirmation')
# def confirmation():
#     # Render a simple confirmation page
#     return "Thank you for your order!"


# NEED TO METHOD = 'POST' THESE ADMIN PAGES
SGT = pytz.timezone('Asia/Singapore')
vehicle_backup_time = []
customer_backup_time = []
logs_backup_time = []

# Session time-out
def admin_session_timeout_check(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        now = datetime.now().timestamp()
        last_activity = session.get('admin_last_activity', None)

        if last_activity and now - last_activity > 15 * 60:  # 15 minutes timeout
            admin_username = session.get('admin_username')
            session.pop('admin_logged_in', None)
            session.pop('admin_username', None)
            session.pop('admin_role', None)
            session.pop('admin_last_activity', None)
            log_event("Logout", f"Inactivity for 15 minutes led to {admin_username}'s session timeout")
            return redirect(url_for('admin_log_in'))

        session['admin_last_activity'] = now
        return f(*args, **kwargs)

    return decorated_function

# Log event function
def log_event(event_type, event_result):
    log = Log(event_type=event_type, event_result=event_result)
    db.session.add(log)
    db.session.commit()


def role_required(*roles):
    def wrapper(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'admin_logged_in' not in session:
                return redirect(url_for('admin_log_in'))

            if session.get('admin_role') not in roles:
                return redirect(url_for('ErrorPage'))

            return f(*args, **kwargs)

        return decorated_function

    return wrapper


def admin_login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_log_in'))  # Redirect to admin login if not logged in
        return f(*args, **kwargs)

    return decorated_function


@app.route('/admin_log_in', methods=['GET', 'POST'])
def admin_log_in():
    form = AdminLoginForm()
    error_message = None

    if form.validate_on_submit():
        username = html.escape(form.username.data)
        password = html.escape(form.password.data)

        if not username.endswith('@ecowheels.com'):
            error_message = "Incorrect Username or Password"
            log_event('Login', f'Failed login attempt for non-existent admin {username}.')
            return render_template('admin/admin_log_in.html', form=form, error_message=error_message)

        admin = db.session.query(Admin).filter(func.binary(Admin.username) == username).first()

        if admin:
            if admin.is_suspended:
                error_message = "Your account is suspended due to too many failed login attempts."
                log_event('Suspension', f'Attempted login by suspended admin {username}.')
                return render_template('admin/admin_log_in.html', form=form, error_message=error_message)

            if admin.check_password(password):
                admin.login_attempts = 0
                db.session.commit()

                session.permanent = True  # Make the session permanent to use PERMANENT_SESSION_LIFETIME
                session['admin_username'] = username
                session['admin_last_activity'] = datetime.now().timestamp()
                session['admin_role'] = admin.role

                if admin.role == 'system':
                    log_event('Login', f'Successful first login for system admin {username}.')
                    if not admin.totp_secret:
                        totp_secret = pyotp.random_base32()
                        admin.totp_secret = totp_secret
                        db.session.commit()

                    session['admin_logged_in'] = False
                    return redirect(url_for('verify_2fa'))

                session['admin_logged_in'] = True

                if admin.role == 'junior':
                    log_event('Login', f'Successful login for junior admin {username}.')
                    return redirect(url_for('sub_dashboard'))
                else:
                    log_event('Login', f'Successful login for {admin.role} admin {username}.')
                    return redirect(url_for('dashboard'))
            else:
                admin.login_attempts += 1
                log_event('Login', f'Failed login attempt for admin {username}. Attempt {admin.login_attempts}')

                if admin.login_attempts >= 3:
                    admin.is_suspended = True
                    error_message = "Your account is suspended due to too many failed login attempts."
                    log_event('Suspension',
                              f'Due to too many failed login attempts, active suspension of admin account {username}.')
                else:
                    error_message = "Incorrect Username or Password"

                db.session.commit()
        else:
            error_message = "Incorrect Username or Password"
            log_event('Login', f'Failed login attempt for non-existent admin {username}.')

    return render_template('admin/admin_log_in.html', form=form, error_message=error_message)


# Route to verify 2FA for system admin
@app.route('/verify_2fa', methods=['GET', 'POST'])
def verify_2fa():
    csrf_token = generate_csrf()
    if 'admin_username' not in session:
        return redirect(url_for('admin_log_in'))

    username = session['admin_username']
    admin = db.session.query(Admin).filter_by(username=username).first()

    if not admin or admin.totp_secret is None:
        return redirect(url_for('admin_log_in'))

    error_message = None
    if request.method == 'POST':
        token = request.form.get('token')
        totp = pyotp.TOTP(admin.totp_secret)
        if totp.verify(token) and is_valid_input(token):
            session['admin_logged_in'] = True
            admin.login_attempts = 0
            admin.is_first_login = False  # Set first login to False after successful 2FA
            db.session.commit()

            log_event('Login', f'Successful 2FA login for system admin {username}.')
            return redirect(url_for('system_dashboard'))
        else:
            error_message = "Invalid 2FA code. Please try again."
            log_event('Login', f'Failed 2FA attempt for system admin {username}.')

    return render_template('admin/system_admin/verify_2fa.html', error_message=error_message, csrf_token=csrf_token, is_first_login=admin.is_first_login)


# Route to serve the QR code image
@app.route('/qr_code')
def qr_code():
    if 'admin_username' not in session:
        return redirect(url_for('admin_log_in'))

    username = session['admin_username']
    admin = db.session.query(Admin).filter_by(username=username).first()

    if not admin or admin.totp_secret is None:
        return redirect(url_for('admin_log_in'))

    totp = pyotp.TOTP(admin.totp_secret)
    uri = totp.provisioning_uri(name=username, issuer_name="EcoWheels")
    img = qrcode.make(uri)
    img = img.resize((200, 200))  # Resize the QR code to 200x200 pixels
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)

    return send_file(buf, mimetype='image/png')


# def is_valid_input(input_str):
#     """
#     Check if the input string contains only allowed characters.
#     """
#     # Define a regular expression to match allowed characters
#     allowed_chars_pattern = re.compile(r'^[\w.@+-]+$')
#     return bool(allowed_chars_pattern.match(input_str))


def save_image_file(form_file):
    if not form_file or not form_file.filename:
        raise BadRequest("No file provided or invalid file.")

    random_hex = secrets.token_hex(8)
    _, f_ext = os.path.splitext(form_file.filename)
    picture_fn = random_hex + f_ext.lower()  # Ensure lowercase extension
    picture_path = os.path.join(current_app.root_path, 'static/vehicle_images', picture_fn)

    # Check if the provided file is a directory before saving
    if hasattr(form_file, 'read'):
        form_file.seek(0)
        form_file.save(picture_path)
    else:
        raise BadRequest("Uploaded data is not a valid file.")

    # Check if the saved file is actually a file
    if os.path.isdir(picture_path):
        os.remove(picture_path)
        raise BadRequest("Uploaded data is a directory, not a file.")

    try:
        Image.open(picture_path).verify()
    except Exception:
        os.remove(picture_path)  # Remove the file if verification fails
        raise ValueError("Invalid image file.")

    return picture_fn


@app.route('/createVehicle', methods=['GET', 'POST'])
@role_required('general')
@admin_login_required
@admin_session_timeout_check
def createVehicle():
    create_vehicle_form = CreateVehicleForm()
    if request.method == 'POST' and create_vehicle_form.validate_on_submit():
        product_id = create_vehicle_form.product_id.data
        brand = create_vehicle_form.brand.data
        model = create_vehicle_form.model.data
        price = create_vehicle_form.price.data
        description = create_vehicle_form.description.data

        if create_vehicle_form.file.data:
            try:
                file = save_image_file(create_vehicle_form.file.data)
            except (BadRequest, ValueError) as e:
                return redirect(url_for('ErrorPage'))
        else:
            file = None

        try:
            new_vehicle = Vehicle(product_id=product_id, brand=brand, model=model, selling_price=price, image=file, description=description)
            db.session.add(new_vehicle)
            db.session.commit()
            log_event('Create Vehicle', f'New vehicle created: {product_id} {brand} {model} by {session["admin_role"]} admin {session["admin_username"]}.')
            return redirect(url_for('MVehicles'))
        except Exception:
            db.session.rollback()

    return render_template('admin/createVehicleForm.html', form=create_vehicle_form)


@app.route('/system_createVehicle', methods=['GET', 'POST'])
@role_required('system')
@admin_login_required
@admin_session_timeout_check
def system_createVehicle():
    create_vehicle_form = CreateVehicleForm()
    if request.method == 'POST' and create_vehicle_form.validate_on_submit():
        product_id = create_vehicle_form.product_id.data
        brand = create_vehicle_form.brand.data
        model = create_vehicle_form.model.data
        price = create_vehicle_form.price.data
        description = create_vehicle_form.description.data

        if create_vehicle_form.file.data:
            try:
                file = save_image_file(create_vehicle_form.file.data)
            except (BadRequest, ValueError) as e:
                return redirect(url_for('ErrorPage'))
        else:
            file = None

        try:
            new_vehicle = Vehicle(product_id=product_id,brand=brand, model=model, selling_price=price, image=file, description=description)
            db.session.add(new_vehicle)
            db.session.commit()
            log_event('Create Vehicle', f'New vehicle created: {product_id} {brand} {model} by {session["admin_role"]} admin {session["admin_username"]}.')
            return redirect(url_for('system_MVehicles'))
        except Exception:
            db.session.rollback()

    return render_template('admin/system_admin/system_createVehicleForm.html', form=create_vehicle_form)


@app.route('/ErrorPage')
def ErrorPage():
    referrer = request.referrer or url_for('home')  # Default to 'home' if no referrer
    return render_template('admin/ErrorPage.html', referrer=referrer)


@app.route('/dashboard', methods=['GET', 'POST'])
@admin_login_required
@admin_session_timeout_check
@role_required('general')
def dashboard():
    admin_username = session.get('admin_username')
    num_customers = db.session.query(User).count()
    num_vehicles = db.session.query(Vehicle).count()
    num_feedbacks = db.session.query(Feedback).count()

    # Calculate the number of each rating type
    good_ratings = db.session.query(Feedback).filter_by(rating='good').count()
    moderate_ratings = db.session.query(Feedback).filter_by(rating='moderate').count()
    bad_ratings = db.session.query(Feedback).filter_by(rating='bad').count()

    # Set security_modal_shown flag if not set
    if 'security_modal_shown' not in session:
        session['security_modal_shown'] = True
        show_security_modal = True
    else:
        show_security_modal = False

    return render_template('admin/dashboard.html', admin_username=admin_username,
                           num_customers=num_customers, num_vehicles=num_vehicles, num_feedbacks=num_feedbacks,
                           show_security_modal=show_security_modal,
                           good_ratings=good_ratings, moderate_ratings=moderate_ratings, bad_ratings=bad_ratings)


@app.route('/system_admin_dashboard', methods=['GET', 'POST'])
@admin_login_required
@admin_session_timeout_check
@role_required('system')
def system_dashboard():
    admin_username = session.get('admin_username')
    num_customers = db.session.query(User).count()
    num_vehicles = db.session.query(Vehicle).count()
    num_admins = db.session.query(Admin).count()
    num_feedbacks = db.session.query(Feedback).count()

    # Calculate the number of each rating type
    good_ratings = db.session.query(Feedback).filter_by(rating='good').count()
    moderate_ratings = db.session.query(Feedback).filter_by(rating='moderate').count()
    bad_ratings = db.session.query(Feedback).filter_by(rating='bad').count()

    # Set security_modal_shown flag if not set
    if 'security_modal_shown' not in session:
        session['security_modal_shown'] = True
        show_security_modal = True
    else:
        show_security_modal = False

    return render_template('admin/system_admin/system_dashboard.html', admin_username=admin_username,
                           num_customers=num_customers, num_vehicles=num_vehicles, num_admins=num_admins, num_feedbacks=num_feedbacks,
                           show_security_modal=show_security_modal,
                           good_ratings=good_ratings, moderate_ratings=moderate_ratings, bad_ratings=bad_ratings)


@app.route('/sub_dashboard', methods=['GET', 'POST'])
@admin_login_required
@admin_session_timeout_check
@role_required('junior')
def sub_dashboard():
    admin_username = session.get('admin_username')
    num_customers = db.session.query(User).count()
    num_vehicles = db.session.query(Vehicle).count()
    num_feedbacks = db.session.query(Feedback).count()

    # Calculate the number of each rating type
    good_ratings = db.session.query(Feedback).filter_by(rating='good').count()
    moderate_ratings = db.session.query(Feedback).filter_by(rating='moderate').count()
    bad_ratings = db.session.query(Feedback).filter_by(rating='bad').count()

    # Set security_modal_shown flag if not set
    if 'security_modal_shown' not in session:
        session['security_modal_shown'] = True
        show_security_modal = True
    else:
        show_security_modal = False

    return render_template('admin/junior_admin/sub_dashboard.html', admin_username=admin_username,
                           num_customers=num_customers, num_vehicles=num_vehicles, num_feedbacks=num_feedbacks,
                           show_security_modal=show_security_modal,
                           good_ratings=good_ratings, moderate_ratings=moderate_ratings, bad_ratings=bad_ratings)


@app.route('/manageCustomers', methods=['GET', 'POST'])
@admin_login_required
@admin_session_timeout_check
@role_required('general')
def MCustomers():
    admin_username = session.get('admin_username')
    csrf_token = generate_csrf()  # Generate CSRF token
    query = db.session.query(User)
    errors = {}

    if request.method == 'POST':
        full_name_filter = request.form.get('full_name_filter')
        username_filter = request.form.get('username_filter')
        email_filter = request.form.get('email_filter')
        phone_number_filter = request.form.get('phone_number_filter')

        if full_name_filter and is_valid_input(full_name_filter):
            query = query.filter(User.full_name.ilike(f"%{full_name_filter}%"))
        elif full_name_filter:
            errors['full_name_filter'] = 'Invalid input for full name filter'

        if username_filter and is_valid_input(username_filter):
            query = query.filter(User.username.ilike(f"%{username_filter}%"))
        elif username_filter:
            errors['username_filter'] = 'Invalid input for username filter'

        if email_filter and is_valid_input(email_filter):
            query = query.filter(User.email.ilike(f"%{email_filter}%"))
        elif email_filter:
            errors['email_filter'] = 'Invalid input for email filter'

        if phone_number_filter and is_valid_input(phone_number_filter):
            query = query.filter(User.phone_number.ilike(f"%{phone_number_filter}%"))
        elif phone_number_filter:
            errors['phone_number_filter'] = 'Invalid input for phone number filter'

    customers = query.all()

    return render_template('admin/manageCustomers.html', admin_username=admin_username, customers=customers,
                           csrf_token=csrf_token, errors=errors)


@app.route('/system_manageCustomers', methods=['GET', 'POST'])
@admin_login_required
@admin_session_timeout_check
@role_required('system')
def system_MCustomers():
    admin_username = session.get('admin_username')
    csrf_token = generate_csrf()  # Generate CSRF token
    query = db.session.query(User)
    errors = {}
    if request.method == 'POST':
        full_name_filter = request.form.get('full_name_filter')
        username_filter = request.form.get('username_filter')
        email_filter = request.form.get('email_filter')
        phone_number_filter = request.form.get('phone_number_filter')

        if full_name_filter and is_valid_input(full_name_filter):
            query = query.filter(User.full_name.ilike(f"%{full_name_filter}%"))
        elif full_name_filter:
            errors['full_name_filter'] = 'Invalid input for full name filter'

        if username_filter and is_valid_input(username_filter):
            query = query.filter(User.username.ilike(f"%{username_filter}%"))
        elif username_filter:
            errors['username_filter'] = 'Invalid input for username filter'

        if email_filter and is_valid_input(email_filter):
            query = query.filter(User.email.ilike(f"%{email_filter}%"))
        elif email_filter:
            errors['email_filter'] = 'Invalid input for email filter'

        if phone_number_filter and is_valid_input(phone_number_filter):
            query = query.filter(User.phone_number.ilike(f"%{phone_number_filter}%"))
        elif phone_number_filter:
            errors['phone_number_filter'] = 'Invalid input for phone number filter'

    customers = query.all()

    return render_template('admin/system_admin/system_manageCustomers.html', admin_username=admin_username,
                           customers=customers, csrf_token=csrf_token, errors=errors,
                           customer_backup_time=customer_backup_time)


@app.route('/sub_manageCustomers', methods=['GET', 'POST'])
@admin_login_required
@admin_session_timeout_check
@role_required('junior')
def sub_MCustomers():
    admin_username = session.get('admin_username')
    csrf_token = generate_csrf()  # Generate CSRF token
    errors = {}

    query = db.session.query(User)
    if request.method == 'POST':
        full_name_filter = request.form.get('full_name_filter')
        username_filter = request.form.get('username_filter')
        email_filter = request.form.get('email_filter')
        phone_number_filter = request.form.get('phone_number_filter')

        if full_name_filter and is_valid_input(full_name_filter):
            query = query.filter(User.full_name.ilike(f"%{full_name_filter}%"))
        elif full_name_filter:
            errors['full_name_filter'] = 'Invalid input for full name filter'

        if username_filter and is_valid_input(username_filter):
            query = query.filter(User.username.ilike(f"%{username_filter}%"))
        elif username_filter:
            errors['username_filter'] = 'Invalid input for username filter'

        if email_filter and is_valid_input(email_filter):
            query = query.filter(User.email.ilike(f"%{email_filter}%"))
        elif email_filter:
            errors['email_filter'] = 'Invalid input for email filter'

        if phone_number_filter and is_valid_input(phone_number_filter):
            query = query.filter(User.phone_number.ilike(f"%{phone_number_filter}%"))
        elif phone_number_filter:
            errors['phone_number_filter'] = 'Invalid input for phone number filter'

    customers = query.all()

    return render_template('admin/junior_admin/sub_manageCustomers.html', admin_username=admin_username,
                           customers=customers, csrf_token=csrf_token, errors=errors)


@app.route('/manageVehicles', methods=['GET', 'POST'])
@admin_login_required
@admin_session_timeout_check
@role_required('general')
def MVehicles():
    admin_username = session.get('admin_username')
    csrf_token = generate_csrf()  # Generate CSRF token
    vehicles = db.session.query(Vehicle).all()
    errors = {}

    if request.method == 'POST':
        brand_filter = request.form.get('brand_filter')
        model_filter = request.form.get('model_filter')
        min_price_filter = request.form.get('min_price_filter')
        max_price_filter = request.form.get('max_price_filter')

        query = db.session.query(Vehicle)
        if brand_filter and is_valid_input(brand_filter):
            query = query.filter(Vehicle.brand.ilike(f"%{brand_filter}%"))
        elif brand_filter:
            errors['brand_filter'] = 'Invalid input for brand filter'

        if model_filter and is_valid_input(model_filter):
            query = query.filter(Vehicle.model.ilike(f"%{model_filter}%"))
        elif model_filter:
            errors['model_filter'] = 'Invalid input for model filter'

        if min_price_filter and is_valid_input(min_price_filter):
            query = query.filter(Vehicle.selling_price >= float(min_price_filter))
        elif min_price_filter:
            errors['min_price_filter'] = 'Invalid input for minimum price filter'

        if max_price_filter and is_valid_input(max_price_filter):
            query = query.filter(Vehicle.selling_price <= float(max_price_filter))
        elif max_price_filter:
            errors['max_price_filter'] = 'Invalid input for maximum price filter'

        vehicles = query.all()

    return render_template('admin/manageVehicles.html', admin_username=admin_username, vehicles=vehicles,
                           csrf_token=csrf_token, errors=errors)


@app.route('/system_manageVehicles', methods=['GET', 'POST'])
@admin_login_required
@admin_session_timeout_check
@role_required('system')
def system_MVehicles():
    admin_username = session.get('admin_username')
    csrf_token = generate_csrf()  # Generate CSRF token
    vehicles = db.session.query(Vehicle).all()
    errors = {}

    if request.method == 'POST':
        brand_filter = request.form.get('brand_filter')
        model_filter = request.form.get('model_filter')
        min_price_filter = request.form.get('min_price_filter')
        max_price_filter = request.form.get('max_price_filter')

        query = db.session.query(Vehicle)
        if brand_filter and is_valid_input(brand_filter):
            query = query.filter(Vehicle.brand.ilike(f"%{brand_filter}%"))
        elif brand_filter:
            errors['brand_filter'] = 'Invalid input for brand filter'

        if model_filter and is_valid_input(model_filter):
            query = query.filter(Vehicle.model.ilike(f"%{model_filter}%"))
        elif model_filter:
            errors['model_filter'] = 'Invalid input for model filter'

        if min_price_filter and is_valid_input(min_price_filter):
            query = query.filter(Vehicle.selling_price >= float(min_price_filter))
        elif min_price_filter:
            errors['min_price_filter'] = 'Invalid input for minimum price filter'

        if max_price_filter and is_valid_input(max_price_filter):
            query = query.filter(Vehicle.selling_price <= float(max_price_filter))
        elif max_price_filter:
            errors['max_price_filter'] = 'Invalid input for maximum price filter'

        vehicles = query.all()

    return render_template('admin/system_admin/system_manageVehicles.html', admin_username=admin_username,
                           vehicles=vehicles, csrf_token=csrf_token, errors=errors,
                           vehicle_backup_time=vehicle_backup_time)


@app.route('/sub_manageVehicles', methods=['GET', 'POST'])
@admin_login_required
@admin_session_timeout_check
@role_required('junior')
def sub_MVehicles():
    admin_username = session.get('admin_username')
    csrf_token = generate_csrf()  # Generate CSRF token
    vehicles = db.session.query(Vehicle).all()
    errors = {}

    if request.method == 'POST':
        brand_filter = request.form.get('brand_filter')
        model_filter = request.form.get('model_filter')
        min_price_filter = request.form.get('min_price_filter')
        max_price_filter = request.form.get('max_price_filter')

        query = db.session.query(Vehicle)
        if brand_filter and is_valid_input(brand_filter):
            query = query.filter(Vehicle.brand.ilike(f"%{brand_filter}%"))
        elif brand_filter:
            errors['brand_filter'] = 'Invalid input for brand filter'

        if model_filter and is_valid_input(model_filter):
            query = query.filter(Vehicle.model.ilike(f"%{model_filter}%"))
        elif model_filter:
            errors['model_filter'] = 'Invalid input for model filter'

        if min_price_filter and is_valid_input(min_price_filter):
            query = query.filter(Vehicle.selling_price >= float(min_price_filter))
        elif min_price_filter:
            errors['min_price_filter'] = 'Invalid input for minimum price filter'

        if max_price_filter and is_valid_input(max_price_filter):
            query = query.filter(Vehicle.selling_price <= float(max_price_filter))
        elif max_price_filter:
            errors['max_price_filter'] = 'Invalid input for maximum price filter'

        vehicles = query.all()

    return render_template('admin/junior_admin/sub_manageVehicles.html', admin_username=admin_username,
                           vehicles=vehicles,
                           csrf_token=csrf_token, errors=errors)


@app.route('/delete_vehicle/<int:id>', methods=['POST'])
@admin_login_required
def delete_vehicle(id):
    vehicle = db.session.query(Vehicle).get(id)
    if vehicle:
        db.session.delete(vehicle)
        db.session.commit()
        log_event('Delete Vehicle', f'Vehicle deleted: {vehicle.brand} {vehicle.model} by {session["admin_role"]} admin {session["admin_username"]}.')
    if session['admin_role'] == 'system':
        return redirect(url_for('system_MVehicles'))
    elif session['admin_role'] == 'general':
        return redirect(url_for('MVehicles'))
    else:
        return redirect(url_for('ErrorPage'))


@app.route('/logs', methods=['GET', 'POST'])
@admin_login_required
@admin_session_timeout_check
@role_required('system')
def system_logs():
    admin_username = session.get('admin_username')
    csrf_token = generate_csrf()  # Generate CSRF token
    errors = {}
    if request.method == 'POST':
        event_type = request.form.get('event_type')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        keyword = request.form.get('keyword')

        query = db.session.query(Log)

        if event_type:
            query = query.filter(Log.event_type == event_type)

        if start_date:
            start_datetime = datetime.strptime(start_date, '%Y-%m-%dT%H:%M')
            query = query.filter(Log.event_time >= start_datetime)

        if end_date:
            end_datetime = datetime.strptime(end_date, '%Y-%m-%dT%H:%M')
            query = query.filter(Log.event_time <= end_datetime)

        if keyword and is_valid_input(keyword):
            query = query.filter(Log.event_result.ilike(f'%{keyword}%'))
        elif keyword:
            errors['keyword'] = 'Invalid input for keyword'

        logs = query.all()
    else:
        logs = db.session.query(Log).all()

    return render_template('admin/system_admin/logs.html', admin_username=admin_username, logs=logs
                           , csrf_token=csrf_token, errors=errors, logs_backup_time=logs_backup_time)


@app.route('/system_manageFeedback', methods=['GET', 'POST'])
@admin_login_required
@admin_session_timeout_check
@role_required('system')
def system_manageFeedback():
    admin_username = session.get('admin_username')
    user_email = session.get('user_email')
    csrf_token = generate_csrf()  # Generate CSRF token
    accepted_ratings = ('good', 'moderate', 'bad')
    errors = {}

    if request.method == 'POST':
        # Get filter criteria from the form
        customer_id = request.form.get('customer_id')
        email = request.form.get('email')
        rating = request.form.get('rating')

        # Build the query with filters
        query = db.session.query(Feedback)

        if customer_id and is_valid_input(customer_id) and customer_id.isdigit():
            query = query.filter(Feedback.user_id == int(customer_id))
        elif customer_id:
            errors['customer_id'] = "Invalid input for customer id"

        if email and is_valid_input(email):
            query = query.filter(Feedback.email.ilike(f"%{email}%"))
        elif email:
            errors['email'] = "Invalid input for email"

        if rating and rating in accepted_ratings:
            query = query.filter(Feedback.rating == rating)
        elif rating:
            errors['rating'] = "Invalid rating"

        # Get all feedback entries based on filters
        feedback_entries = query.all()
    else:
        # Query the Feedback table to get all feedback entries
        feedback_entries = db.session.query(Feedback).all()

    # Render the template with the feedback entries
    return render_template('admin/system_admin/system_manageFeedback.html', admin_username=admin_username,
                           feedback_entries=feedback_entries, csrf_token=csrf_token, errors=errors)



@app.route('/system_manageAdmin', methods=['GET', 'POST'])
@admin_login_required
@admin_session_timeout_check
@role_required('system')
def system_manageAdmin():
    admin_username = session.get('admin_username')
    csrf_token = generate_csrf()  # Generate CSRF token
    errors = {}
    if 'admin_logged_in' in session and session['admin_role'] == 'system':
        admins = db.session.query(Admin).all()

        if request.method == 'POST':
            admin_id = request.form.get('admin_id')
            admin_to_unsuspend = db.session.query(Admin).filter_by(id=admin_id).first()
            if admin_to_unsuspend:
                admin_to_unsuspend.is_suspended = False
                db.session.commit()
                return redirect(url_for('system_manageAdmin'))

        return render_template('admin/system_admin/system_manageAdmins.html',
                               admins=admins,
                               admin_username=admin_username,
                               errors=errors,
                               csrf_token=csrf_token)
    else:
        return redirect(url_for('admin_log_in'))


@app.route('/unsuspend_admin', methods=['POST'])
@admin_login_required
@role_required('system')
def unsuspend_admin():
    admin_username = session.get('admin_username')
    admin_id = request.form.get('admin_id')
    admin_password = request.form.get('admin_password')

    current_admin_username = session.get('admin_username')
    current_admin = db.session.query(Admin).filter_by(username=current_admin_username).first()
    if not admin_password or not is_valid_input(admin_password) or not current_admin.check_password(admin_password):
        log_event('Unsuspended', f'System admin {admin_username} has failed to unsuspend an admin')
        return redirect(url_for('system_manageAdmin'))

    if current_admin and current_admin.check_password(admin_password):
        admin = db.session.query(Admin).filter_by(id=admin_id).first()
        if admin:
            admin.is_suspended = False
            admin.login_attempts = 0  # Reset login attempts
            db.session.commit()
            log_event('Unsuspended',
                      f'System admin {current_admin_username} has successfully unsuspended {admin.role} admin {admin.username}')
        else:
            flash('Admin not found', 'error')
    else:
        flash('Incorrect password', 'error')

    return redirect(url_for('system_manageAdmin'))


@app.route('/admin_logout')
def admin_logout():
    if 'admin_logged_in' in session:
        admin_username = session.get('admin_username')
        if session['admin_role'] == 'system':
            log_event('Logout', f'Successfully logged out system admin {admin_username}')
        elif session['admin_role'] == 'junior':
            log_event('Logout', f'Successfully logged out junior admin {admin_username}')
        else:
            log_event('Logout', f'Successfully logged out admin {admin_username}')
        session.pop('admin_logged_in', None)
        session.pop('admin_username', None)
        session.pop('admin_role', None)
        session.pop('admin_last_activity', None)
        session.pop('security_modal_shown', None)
        return redirect(url_for('admin_log_in'))
    else:
        return "Admin is not logged in."


@app.route('/verify_backup', methods=['POST'])
@admin_login_required
@role_required('system')
def verify_backup():
    current_admin_username = session.get('admin_username')
    admin_password = request.form.get('password')
    role = session.get('admin_role')
    totp_code = request.form.get('totp_code')

    if not admin_password or not is_valid_input(admin_password):
        return jsonify({"status": False, "message": "Password is required and no special characters"}), 400

    current_admin = db.session.query(Admin).filter_by(username=current_admin_username).first()

    if current_admin and current_admin.check_password(admin_password) and is_valid_input(admin_password):
        totp = pyotp.TOTP(current_admin.totp_secret)
        if totp.verify(totp_code) and is_valid_input(totp_code):
            return jsonify({"status": True})
    log_event('Backup', f'Attempted Backed up of database by {role} admin {current_admin_username}')
    return jsonify({"status": False, "message": "Invalid credentials or verification failed"}), 400


@app.route('/backup_vehicles', methods=['GET'])
@admin_login_required
@role_required('system')
def backup_vehicles():
    vehicles = db.session.query(Vehicle).all()
    backup_time = datetime.now(SGT).strftime('%d-%m-%Y, %I:%M:%S %p')  # Use the new format
    vehicle_backup_time.append(backup_time)
    admin_username = session.get('admin_username')
    log_event('Backup', f'Backed up Vehicles database by {session["admin_role"]} admin {admin_username}')

    # Create a DataFrame
    vehicle_data = []
    for vehicle in vehicles:
        vehicle_data.append({
            'Vehicle ID': vehicle.idvehicles,
            'Brand': vehicle.brand,
            'Model': vehicle.model,
            'Selling Price': vehicle.selling_price,
            'Image': vehicle.image,
            'Description': vehicle.description
        })

    df = pd.DataFrame(vehicle_data)

    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicles"

    # Write the DataFrame to the Excel sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Add images to the Excel sheet
    row_index = 2  # Start from the second row
    for vehicle in vehicles:
        if vehicle.image:
            img_path = os.path.join(app.static_folder, 'vehicle_images', vehicle.image)
            if os.path.exists(img_path):
                img = XLImage(img_path)
                img.height = 100  # Set image height
                img.width = 100  # Set image width
                ws.add_image(img, f'E{row_index}')  # Place the image in the correct cell
        row_index += 1  # Increment the row index

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the user
    return send_file(output, download_name='backupVehicle.xlsx', as_attachment=True)


@app.route('/backup_customers', methods=['GET'])
@admin_login_required
@role_required('system')
def backup_customers():
    # Query all customers from the database
    customers = db.session.query(User).all()
    backup_time = datetime.now(SGT).strftime('%d-%m-%Y, %I:%M:%S %p')  # Use the new format
    customer_backup_time.append(backup_time)
    admin_username = session.get('admin_username')
    log_event('Backup', f'Backed up Customers database by {session["admin_role"]} admin {admin_username}')

    # Create a DataFrame with relevant customer data
    customer_data = []
    for customer in customers:
        customer_data.append({
            'Customer ID': customer.id,
            'Full Name': customer.full_name,
            'Username': customer.username,
            'Email': customer.email,
            'Phone Number': customer.phone_number
        })

    df = pd.DataFrame(customer_data)

    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    # Write the DataFrame to the Excel sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Save the Excel file to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the user
    return send_file(output, download_name='backupCustomers.xlsx', as_attachment=True)


@app.route('/backup_logs', methods=['GET'])
@admin_login_required
@role_required('system')
def backup_logs():
    logs = db.session.query(Log).all()
    backup_time = datetime.now(SGT).strftime('%d-%m-%Y, %I:%M:%S %p')  # Use the new format
    logs_backup_time.append(backup_time)
    admin_username = session.get('admin_username')
    log_event('Backup', f'Backed up Logs database by {session["admin_role"]} admin {admin_username}')

    # Create a DataFrame
    log_data = []
    for log in logs:
        log_data.append({
            'Log ID': log.id,
            'Event Type': log.event_type,
            'Event Time': log.event_time.strftime('%Y-%m-%d %H:%M:%S'),
            'Event Result': log.event_result
        })

    df = pd.DataFrame(log_data)

    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs"

    # Define color fills based on logs.html CSS
    colors = {
        'Success': '73de8c',       # Light green (log-success)
        'Failed': 'f77674',        # Red (log-failure)
        'created': 'f5c842',       # Orange-yellow (log-creation)
        'deleted': 'faa441',       # Bright orange (log-deletion)
        'suspension': 'fa7364',    # Dark red (log-suspend)
        'Attempted': 'e65545',     # Darker red (log-attempt)
        'unsuspend': '5dc44d',     # Dark green (log-unsuspend)
        'customer': '7ddce8',      # Light blue (log-unlock)
        'Backup': 'c2c2c2'         # Light gray (Backup)
    }

    # Apply header styles
    header_fill = PatternFill(start_color='d9ead3', end_color='d9ead3', fill_type='solid')
    for c in range(1, len(df.columns) + 1):
        ws.cell(row=1, column=c).fill = header_fill

    # Write the DataFrame to the Excel sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Apply color based on event result
            if r_idx > 1:  # Skip header row
                event_result = df.iloc[r_idx - 2]['Event Result']
                for key in colors:
                    if key.lower() in event_result.lower():
                        color = colors[key]
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        break
                else:
                    # Default to white if no matching color found
                    cell.fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the user
    return send_file(output, download_name='backupLogs.xlsx', as_attachment=True)


@app.route('/unlock_customer', methods=['POST'])
@admin_login_required
@role_required('system')
def unlock_customer():
    current_admin_username = session.get('admin_username')
    customer_id = request.form.get('customer_id')
    admin_password = request.form.get('admin_password')

    current_admin = db.session.query(Admin).filter_by(username=current_admin_username).first()

    if not admin_password or not is_valid_input(admin_password) or not current_admin.check_password(admin_password):
        log_event('Unlock', f'System admin {current_admin_username} has failed to unlock a customer')
        return redirect(url_for('system_MCustomers'))

    if current_admin and current_admin.check_password(admin_password):
        customer = db.session.query(User).filter_by(id=customer_id).first()
        if customer:
            customer.lockout_count = 0
            customer.lockout_until = None
            customer.failed_attempts = 0
            db.session.commit()
            log_event('Unlock', f'System admin {current_admin_username} unlocked customer {customer.email}')
        else:
            flash('Customer not found', 'error')
    else:
        flash('Incorrect password', 'error')

    return redirect(url_for('system_MCustomers'))

def is_valid_email(email):
    regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
    return re.match(regex, email) is not None

@app.route('/purchaseHistory', methods=['POST'])
@admin_login_required
@admin_session_timeout_check
def purchase_history():
    try:
        data = request.get_json()
        email = data.get('email')

        # Validate email format
        if not is_valid_email(email):
            return jsonify({'error': 'Invalid email format'}), 400

        # Query the PurchasedItem table for the given customer email
        purchases = db.session.query(Product).filter_by(email=email).all()

        # Prepare the data to be sent as JSON
        purchase_data = [
            {
                'product_name': purchase.product_name,
                'price': float(purchase.price)
            }
            for purchase in purchases
        ]

        # Get the total count of purchased items
        purchase_count = len(purchases)

        return jsonify({'purchases': purchase_data, 'count': purchase_count})
    except Exception as e:
        app.logger.error(f"Error fetching purchase history: {e}")
        return jsonify({'error': 'An error occurred while fetching purchase history'}), 500


SUSPICIOUS_PATTERNS = [
    # Injection
    r"union\s+select",
    r"select\s+\*+\s+from",
    r"drop\s+table",
    r"insert\s+into",
    r"update\s+set",
    r"select\s+from\s+information_schema.tables",
    r"union\s+all\s+select",
    r"cmd\s*=\s*[\S]+",
    r"wget",
    r"curl",
    r"nc\s+-e",
    r"powershell\s+",
    r"eval\(",

    # Broken Access Control
    r"admin\s+access",
    r"unauthorized\s+access",
    r"access\s+denied",

    # Cryptographic Failures
    r"api\s+key",
    r"password\s*=",
    r"secret\s*=",
    r"token\s*=",
    r"ssn\s*=",

    # Security Misconfiguration
    r"server\s+status",
    r"debug\s+mode",
    r"config\s+file",

    # Vulnerable and Outdated Components
    r"vulnerable\s+component",
    r"known\s+vulnerability",
    r"component\s+version",

    # Identification and Authentication Failures
    r"login\s*failed",
    r"invalid\s+username",
    r"invalid\s+password",
    r"forgot\s+password",

    # Security Logging and Monitoring Failures
    r"error\s+log",
    r"monitoring\s+failed",
    r"access\s+log",

    # Server-Side Request Forgery (SSRF)
    r"localhost",
    r"127\.0\.0\.1",

    # XSS Patterns
    r"script\s+src=",
    r"<img\s+src=",
    r"data:text/html",
]


@app.before_request
def log_and_check_request():
    ip_src = request.remote_addr
    query_params = request.args.to_dict() 
    form_data = request.form.to_dict()

    # Decodes the URL-encoded values in the query params + form data
    decoded_params = {key: unquote(value) for key, value in query_params.items()}
    decoded_form_data = {key: unquote(value) for key, value in form_data.items()}

    request_data = f"Query Params: {decoded_params}, Form Data: {decoded_form_data}"

    if decoded_params or decoded_form_data:
        app.logger.info(f"Decoded Request Data: {request_data}")

    for pattern in SUSPICIOUS_PATTERNS:
        if re.search(pattern, request_data, re.IGNORECASE):
            app.logger.warning(f"Suspicious activity detected from {ip_src}: pattern '{pattern}' detected")
            abort(403)


if __name__ == '__main__':
    app.run(debug=True, port = 5000)
